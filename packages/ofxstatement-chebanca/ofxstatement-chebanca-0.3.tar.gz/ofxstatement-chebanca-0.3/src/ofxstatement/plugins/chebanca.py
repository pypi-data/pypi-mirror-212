import logging

from decimal import Decimal
from enum import Enum
from typing import Any, Iterable, Optional

from ofxstatement.plugin import Plugin
from ofxstatement.parser import StatementParser
from ofxstatement.statement import (
    BankAccount,
    Currency,
    Statement,
    StatementLine,
    generate_transaction_id,
    recalculate_balance,
)

from openpyxl import load_workbook
from openpyxl.cell import Cell

logging.basicConfig(level=logging.DEBUG)
logger = logging.getLogger("CheBanca")

TYPE_MAPPING = {
    "Accrediti diversi": "CREDIT",
    "Addebito Canone": "FEE",
    "Addebito canone": "FEE",
    "Addebito Carta": "PAYMENT",
    "Addebito SDD": "DIRECTDEBIT",
    "Addebito SDD": "DIRECTDEBIT",
    "Addebito/Accredito competenze": "INT",
    "Bancomat": "ATM",
    "Bonif. v/fav.": "XFER",
    "Bonifico a vostro favore per ordine e conto": "XFER",
    "Bonifico dall'estero": "XFER",
    "Bonifico": "XFER",
    "Carta Credito.": "PAYMENT",
    "cont. ATM": "ATM",
    "Delega Unica": "PAYMENT",
    "Disposizione di pagamento": "XFER",
    "Disposizione": "XFER",
    "Giroconto": "XFER",
    "Pagam. POS": "POS",
    "Pagamenti diversi": "PAYMENT",
    "Pagamento imposte Delega Unificata": "PAYMENT",
    "Pagamento imposte e tasse": "FEE",
    "Pagamento per utilizzo carta di credito": "PAYMENT",
    "Pagamento tramite POS": "POS",
    "Prelievo Bancomat altri Istituti": "ATM",
    "Prelievo Bancomat": "ATM",
    "Stipendio": "XFER",
    "Storno disposizione di pagamento": "XFER",
}


class Fields(Enum):
    DATE = "Data contabile"
    USER_DATE = "Data valuta"
    TYPE = "Tipologia"
    IN = "Entrate"
    OUT = "Uscite"
    CURRENCY = "Divisa"


class CheBancaParser(StatementParser[str]):
    date_format = "%d/%m/%Y"

    def __init__(self, filename: str) -> None:
        super().__init__()
        self.filename = filename

        logging.debug(f"Loading {self.filename}")
        self._ws = load_workbook(self.filename).active
        self._fields_to_row = {}

    def parse(self) -> Statement:
        found = False

        fields_values = [f.value.lower() for f in Fields]
        for row in self._ws:
            for cell in row:
                if isinstance(cell.value, str) and (
                    cell.value.lower() in fields_values
                ):
                    start_row = cell.row
                    start_column = cell.col_idx - 1
                    found = True
                    break

            if found:
                break

        if not found:
            raise ValueError("No 'Data contabile' cell found")

        logging.debug(
            "Statement table start cell found at "
            f"{self._ws[start_row][start_column].coordinate}"
        )

        for field in Fields:
            for cell in self._ws[start_row][start_column:]:
                if cell.value == field.value:
                    self._fields_to_row[field] = cell.col_idx - start_column - 1
                    break

        logging.debug(f"Statement table mapping are {self._fields_to_row}")

        if not [Fields.DATE, Fields.USER_DATE] & self._fields_to_row.keys():
            raise ValueError("No date column found")

        if not [Fields.IN, Fields.OUT] & self._fields_to_row.keys():
            raise ValueError("No amount column found")

        if not Fields.TYPE in self._fields_to_row.keys():
            raise ValueError("No type column")

        self._start_row = start_row + 1
        self._start_column = start_column

        for row in self._ws.iter_rows(min_row=1, max_row=self._start_row):
            for cell in row:
                if isinstance(cell.value, str) and "IBAN:" in cell.value:
                    self.statement.account_id = row[cell.col_idx].value

                if (
                    isinstance(cell.value, str)
                    and "Saldo disponibile:" in cell.value
                    and not self.statement.end_balance
                ):
                    self.statement.end_balance = self.parse_value(
                        row[cell.col_idx].value, "amount"
                    )

                if isinstance(cell.value, str) and "Saldo contabile:" in cell.value:
                    self.statement.end_balance = self.parse_value(
                        row[cell.col_idx].value, "amount"
                    )

                if isinstance(cell.value, str) and "Divisa:" in cell.value:
                    self.statement.currency = (
                        row[cell.col_idx].value.strip()
                        if row[cell.col_idx].value
                        else None
                    )

                if isinstance(cell.value, str) and "PERIODO:" in cell.value:
                    splits = cell.value.split(": ", 1)[-1].split(" fino al ")
                    if len(splits) > 1:
                        if splits[0].startswith("dal "):
                            self.statement.start_date = self.parse_value(
                                splits[0][4:], "date"
                            )
                            self.statement.end_date = self.parse_value(
                                splits[-1], "date"
                            )

        if self.statement.account_id:
            logger.debug(f"Account ID: {self.statement.account_id}")

        if self.statement.currency:
            logger.debug(f"Currency: {self.statement.currency}")

        self._bank_account = BankAccount(
            bank_id="MICSITM1XXX", acct_id=self.statement.account_id
        )

        statement = super().parse()

        if statement.end_balance:
            total_amount = sum(
                [sl.amount for sl in statement.lines if sl.amount is not None], 0
            )
            statement.start_balance = statement.end_balance - total_amount

        recalculate_balance(statement)

        if self.statement.start_balance:
            logger.debug(
                f"Start balance: {self.statement.start_balance:0.2f} at {self.statement.start_date}"
            )

        if self.statement.end_balance:
            logger.debug(
                f"End balance: {self.statement.end_balance:0.2f} at {self.statement.end_date}"
            )

        return statement

    def split_records(self) -> Iterable[Iterable[Cell]]:

        cells = []
        row = self._start_row
        while True:
            line_contents = self._ws[row][self._start_column :]

            if not any(cell.value for cell in line_contents):
                break

            cells.append(line_contents)
            row += 1

        return cells

    def get_field_record(self, cells: Iterable[Cell], field: Fields) -> Any:
        if field not in self._fields_to_row.keys():
            return None

        return cells[self._fields_to_row[field]].value

    def strip_spaces(self, string: str):
        return " ".join(string.strip().split())

    def parse_value(self, value: Optional[str], field: str) -> Any:
        if field == "trntype":
            native_type = value.split(" - ", 1)[0].strip()
            trntype = TYPE_MAPPING.get(native_type)

            if not trntype:
                logger.warning(f"Mapping not found for {value}")
                return "OTHER"

            return trntype

        elif field == "memo":
            try:
                return self.strip_spaces(value.split(" - ", 1)[1])
            except:
                pass

        if field == "amount" and isinstance(value, float):
            return Decimal(value)

        return super().parse_value(value, field)

    def parse_record(self, cells: Iterable[Cell]) -> StatementLine:
        stat_line = StatementLine(
            date=self.parse_value(
                self.get_field_record(cells, Fields.DATE)
                or self.get_field_record(cells, Fields.USER_DATE),
                "date",
            ),
            memo=self.parse_value(self.get_field_record(cells, Fields.TYPE), "memo"),
            amount=self.parse_value(
                self.get_field_record(cells, Fields.IN)
                or self.get_field_record(cells, Fields.OUT),
                "amount",
            ),
        )

        stat_line.date_user = self.parse_value(
            self.get_field_record(cells, Fields.USER_DATE), "date"
        )
        stat_line.trntype = self.parse_value(
            self.get_field_record(cells, Fields.TYPE), "trntype"
        )

        currency = self.parse_value(
            self.get_field_record(cells, Fields.CURRENCY), "currency"
        )

        if currency:
            stat_line.currency = Currency(symbol=currency)

        if stat_line.memo and stat_line.memo.startswith("RIF:"):
            stat_line.refnum = stat_line.memo.split("RIF:", 1)[1].split(".")[0]

        stat_line.bank_account_to = self._bank_account
        stat_line.id = generate_transaction_id(stat_line)

        logging.debug(stat_line)
        stat_line.assert_valid()

        return stat_line


class CheBancaPlugin(Plugin):
    """CheBanca! parser"""

    def get_parser(self, filename: str) -> CheBancaParser:
        return CheBancaParser(filename)
