from datetime import date, datetime, timezone
from typing import Dict, Any, List, Tuple, Set

from spreadsheet_migrator.spreadsheet_migrator_lib.testy_creator import TestyCreator
from spreadsheet_migrator.spreadsheet_migrator_lib.testy_exception import TestyException

from tests_description.models import TestCase
from functools import reduce
from django.db.models import Q
from core.models import Project
import pytz
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet


class Parser:

    def __init__(self, request_data, testy_creator, config):
        self.excel_data_ws: Worksheet = openpyxl.load_workbook(request_data["file"]).active
        self.request_data = request_data
        self.testy_creator: TestyCreator = testy_creator
        self.config: Dict[str, Dict[str, int]] = config
        if len(request_data['projectId']) > 0:
            self.project = Project.objects.get(id=int(request_data['projectId']))
        else:
            raise TestyException("Project id value is missing.")
        if len(request_data['uuid']) > 0:
            self.uuid = request_data['uuid']
        else:
            raise TestyException("UUID is missing, the report cannot be generated without it.")

    def validate_cell(self, i, config: Dict[str, int], field: str):
        """
        The function checks whether there is a value in the table cell.
        """
        return self.excel_data_ws[i][config[field]].value is not None

    def get_cell_value(self, row, column):
        """
        The function get value in cell.
        """
        return self.excel_data_ws[row][column].value

    def get_case(self, i, config_case: Dict[str, int]):
        """
        The function parses the case data on the line number i.
        """
        case = {}
        if config_case.get("name") is not None and config_case.get("scenario") is not None:
            if self.validate_cell(i, config_case, "name") and self.validate_cell(i, config_case, "scenario"):
                case = {"name": self.get_cell_value(i, config_case.get("name")),
                        "project": self.project.id,
                        "scenario": self.get_cell_value(i, config_case.get("scenario"))}
                if config_case.get("description") is not None and self.validate_cell(i, config_case,
                                                                                     "description"):
                    case["description"] = self.get_cell_value(i, config_case.get("description"))
                if config_case.get("setup") is not None and self.validate_cell(i, config_case,
                                                                               "setup"):
                    case["setup"] = self.get_cell_value(i, config_case.get("setup"))
                if config_case.get("teardown") is not None and self.validate_cell(i, config_case,
                                                                                  "teardown"):
                    case["teardown"] = self.get_cell_value(i, config_case.get("teardown"))
                if config_case.get("estimate") is not None and self.validate_cell(i,
                                                                                  config_case,
                                                                                  "estimate"
                                                                                  ):
                    case["estimate"] = self.get_cell_value(i, config_case.get("estimate"))
            else:
                missing_columns = []
                if not self.validate_cell(i, config_case, "name"):
                    missing_columns.append("name")
                if not self.validate_cell(i, config_case, "scenario"):
                    missing_columns.append("scenario")
                self.testy_creator.cases_logs.lack_data.append(
                    {"row": i, "columns": missing_columns})
        return case

    def get_parameters(self, i, config_parameter: Dict[str, int]):
        """
        The function parses the parameter data on the line number i.
        """
        parameters: List[Dict] = []
        if config_parameter.get("group_data") is not None:
            if self.validate_cell(i, config_parameter, "group_data"):
                group_of_parameters = self.get_cell_value(i, config_parameter.get("group_data")).strip().rstrip(
                    ";").split(";\n")
                for group_name_datas in group_of_parameters:
                    group_name, str_datas = group_name_datas.strip().split(":")
                    datas = str_datas.strip().split(";")
                    for data in datas:
                        parameters.append({"project": self.project.id, "group_name": group_name.strip(),
                                           "data": data.strip()})
            else:
                self.testy_creator.parameters_logs.lack_data.append({"row": i, "columns": ["group_data"]})
        return parameters

    def get_plan(self, i, config_plan: Dict[str, int]) -> Dict[str, Any]:
        """
        The function parses the plan data on the line number i.
        """
        plan = {}
        if config_plan.get("name") is not None:
            datetime_now = datetime.now(timezone.utc)
            missing_columns = []
            if config_plan.get("started_at") is None:
                plan["started_at"] = datetime_now
            elif not self.validate_cell(i, config_plan, "started_at"):
                missing_columns.append("started_at")
                plan["started_at"] = datetime_now
            else:
                plan["started_at"] = self.get_cell_value(i, config_plan.get("started_at")).replace(tzinfo=pytz.utc)
            if config_plan.get("due_date") is None:
                plan["due_date"] = datetime_now
            elif not self.validate_cell(i, config_plan, "due_date"):
                missing_columns.append("due_date")
                plan["due_date"] = datetime_now
            else:
                plan["due_date"] = self.get_cell_value(i, config_plan.get("due_date")).replace(tzinfo=pytz.utc)
            if self.validate_cell(i, config_plan, "name"):
                plan["name"] = self.get_cell_value(i, config_plan.get("name"))
                if config_plan.get("description") is not None and self.validate_cell(i, config_plan, "description"):
                    plan["description"] = self.get_cell_value(i, config_plan.get("description"))
            else:
                if not self.validate_cell(i, config_plan, "name"):
                    missing_columns.append("name")
                self.testy_creator.plans_logs.lack_data.append({"row": i, "columns": missing_columns})
        return plan

    def parse_datas_without_suites(self) -> List[Tuple[List[Dict], Dict]]:
        """
        The function receives data about parameters and plans, is called if you do not need to create suites.
        """
        parameters_plan: List[Tuple[List[Dict], Dict]] = []
        for i in range(2, self.excel_data_ws.max_row + 1):
            parameters: List[Dict] = []
            plan: Dict[str, Any] = {}
            if self.config.get("parameter") is not None:
                parameters = self.get_parameters(i, self.config.get("parameter"))
            if self.config.get("plan") is not None:
                plan = self.get_plan(i, self.config.get("plan"))
            parameters_plan.append((parameters, plan))
        return parameters_plan

    def parse_datas_with_suites(self) -> Dict[Tuple[str, str], List[Tuple[Dict, List[Dict], Dict]]]:
        """
        The function receives data about suites, test-cases, parameters and test-plans,
        these data are combined by the name of the suite.

        Examples:
        suite = ("Authorization", "Check authorization")
        case = {"name": "log in", ...}
        parameters = [{"group_name": "OS", ...}, {"group_name": "BROWSER", ...}]
        plan = {"name": "Release", ...}
        suite_case_parameters_plan:
            Dict[Tuple[str, str], List[Tuple[Dict, List[Dict], Dict]]]
                                                                = {suite: [(case, parameters, plan_name)]}
        """
        suite_case_parameters_plan: Dict[Tuple[str, str], List[Tuple[Dict, List[Dict], Dict]]] = {}
        for i in range(2, self.excel_data_ws.max_row + 1):
            parameters: List[Dict] = []
            plan: Dict[str, Any] = {}
            if self.config.get("parameter") is not None:
                parameters: List[Dict] = self.get_parameters(i, self.config.get("parameter"))
            if self.config.get("plan") is not None:
                plan = self.get_plan(i, self.config.get("plan"))
            if self.validate_cell(i, self.config.get("suite"), "name"):
                case = {}
                if self.config.get("case") is not None:
                    case: Dict[str, Any] = self.get_case(i, self.config.get("case"))
                suite_name = self.get_cell_value(i, self.config.get("suite").get("name"))
                suite = {"name": suite_name}
                if self.config.get("suite").get("description") is not None and \
                        self.validate_cell(i, self.config.get("suite"), "description"):
                    suite["description"] = self.get_cell_value(i, self.config.get("suite").get("description"))
                if (suite.get("name"), suite.get("description", "")) in suite_case_parameters_plan:
                    list_of_case_parameters_plan = suite_case_parameters_plan[
                        (suite.get("name"), suite.get("description", ""))]
                    list_of_case_parameters_plan.append((case, parameters, plan))
                else:
                    suite_case_parameters_plan[(suite.get("name"), suite.get("description", ""))] = (
                        [(case, parameters, plan)])
            else:
                self.testy_creator.suites_logs.lack_data.append({"row": i, "columns": ["name"]})
                self.testy_creator.create_datas_without_suite([(parameters, plan)], self.project)
        return suite_case_parameters_plan

    @staticmethod
    def union_cases_by_equal_parameters(
            dict_idcase_plan_idparameters: Dict[
                int, Dict[Tuple[str, str, str, str], Set[int]]]) -> \
            Dict[Tuple[str, str, str, str], Tuple[Set, Set]]:
        """
        The function combines test-cases with the same parameters
        and the name of the test plan, for further creation
        of a test plan with the found test-cases and parameters.

        Examples:
        plan = ("Release", started_at, due_date)
        idcases = {1, 2, 3}
        idparameters = {1, 2, 3}
        plan_idcases_idparameters: Dict[Tuple[str, str, str], Tuple[Set, Set]] = {plan: (idcases, idparameters)}
        """
        plan_idcases_idparameters: Dict[Tuple[str, str, str, str], Tuple[Set, Set]] = {}
        for case_id, dict_plan_parameters in dict_idcase_plan_idparameters.items():
            for plan, parameters in dict_plan_parameters.items():
                if plan in plan_idcases_idparameters:
                    cases_for_plans, parameters_for_plans = plan_idcases_idparameters[plan]
                    if parameters == parameters_for_plans:
                        cases_for_plans.add(case_id)
                    else:
                        arr_cases = TestCase.objects.filter(
                            reduce(lambda x, y: x | y, [Q(id=case_id) for case_id in cases_for_plans]))
                        one_case = TestCase.objects.get(id=case_id)
                        raise TestyException(
                            f"You cannot choose test cases "
                            f"{[(case_from_arr.id, case_from_arr.name) for case_from_arr in arr_cases]} "
                            f"and {(one_case.id, one_case.name)} "
                            f"for creation tests within the same test plan, since "
                            f"they have conflicting parameters")
                else:
                    plan_idcases_idparameters[plan] = ({case_id}, parameters)
        return plan_idcases_idparameters
