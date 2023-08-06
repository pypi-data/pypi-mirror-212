from openpyxl import load_workbook

ARR = tuple[int,int] # 棋子位置数组

class MapViewer:
    # 解析位置条件
    @staticmethod
    def parse_location(data:str):
        """
        X[1]&Y[2]&P[2,3]&T[7]
        ->
        [('X',1),('Y',2),('P',2,3),('T',7)]
        """
        data = data.split("&")
        loc = list[tuple]()
        for i in data:
            if not i:
                continue
            command = i.split("[")[0]
            args = [int(j) for j in i[len(command) + 1:-1].split("|")]
            loc.append(tuple([command] + args))
        return loc

    # 解析移动规则
    @staticmethod
    def parse_move(data:str):
        """
        1(Y[1|2]),2(Y[3|4]),3
        ->
        [(1,[('Y',1,2)]),(2,[('Y',3,4)]),(3,)]
        """
        data = data.split(",")
        move = list[tuple]()
        for i in data:
            if not i:
                continue
            if "(" in i and i.endswith(")"):
                move.append((
                    int(i.split("(")[0]),
                    MapViewer.parse_location(i.split("(")[1][:-1])
                ))
            else:
                move.append((int(i),))
        return move

    @staticmethod
    def view(filename:str):
        wb = load_workbook(filename)
        # 处理棋子
        chesses_sheet = wb.get_sheet_by_name("chesses")
        chesses = []
        title = chesses_sheet[1]
        for i in chesses_sheet[2:chesses_sheet.max_row]:
            name = str(i[1].value).strip('"')
            belong = int(i[2].value)
            is_captain = i[3].value == "c"
            move = [MapViewer.parse_move(j) for j in str(i[4].value).split(";")]
            tran_con = MapViewer.parse_location(i[5].value) if i[5].value else []
            tran_move = [MapViewer.parse_move(j) for j in str(i[6].value).split(";")]
            attr = dict([(title[j + 7].value,k.value) for j,k in enumerate(i[7:]) if k.value])
            if len(move) < 2 or len(tran_move) < 2:
                raise TypeError
            chesses.append([name,belong,is_captain,move,tran_con,tran_move,attr])
        # 处理地图
        map_sheet = wb.get_sheet_by_name("map")
        map = []
        rl = int(map_sheet[1][1].value)
        cl = int(map_sheet[1][3].value)
        for i in range(rl):
            map.append([])
            for j in range(cl):
                val = map_sheet[i + 2][j]
                map[-1].append(int(val.value) if val.value else None)
        # 处理特殊规则
        rules_sheet = wb.get_sheet_by_name("rules")
        rules = {}
        rules["tran"] = rules_sheet[2][2].value == "c" # 启用升变
        rules["back"] = rules_sheet[3][2].value == "c" # 启用悔棋
        rules["restrict_move_ne"] = rules_sheet[4][2].value == "c" # 限制连续3步以上移动中立棋子
        return (chesses,map,rules)
