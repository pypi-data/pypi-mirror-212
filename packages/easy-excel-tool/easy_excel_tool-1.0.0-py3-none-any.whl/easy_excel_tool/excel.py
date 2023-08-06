"""
@author: hanxinkong
@software: PyCharm
@file: excel_tools.py
@time: 2022/8/28 23:10
"""
import os

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows


class MyException(Exception):
    def __init__(self, *args):
        self.args = args


class CreateError(MyException):
    def __init__(self, code=100, message='Create Exception', args=('Create Exception',)):
        self.args = args
        self.message = message
        self.code = code

    def __str__(self):
        return self.message


class OpenError(MyException):
    def __init__(self, code=101, message='Open Exception', args=('Open Exception',)):
        self.args = args
        self.message = message
        self.code = code

    def __str__(self):
        return self.message


class ParamError(MyException):
    def __init__(self, code=102, message='Parameter exception', args=('Parameter exception',)):
        self.args = args
        self.message = message
        self.code = code

    def __str__(self):
        return self.message


class ParamTypeError(MyException):
    def __init__(self, code=103, message='Wrong parameter type', args=('Wrong parameter type',)):
        self.args = args
        self.message = message
        self.code = code

    def __str__(self):
        return self.message


class FileExistenceError(MyException):
    def __init__(self, code=104, message='The file already exists', args=('The file already exists',)):
        self.args = args
        self.message = message
        self.code = code

    def __str__(self):
        return self.message


class Excel(object):
    def __init__(self, *args, **kwargs):
        """
        This is a very convenient tool for operating excel.
        After the secondary encapsulation of excel operation by Pandas,
        The purpose of this tool is to make it easier without focusing on the process.
        :param args:
        :param kwargs:
        """
        self.__wb = None
        self.file = args[0]

    def create_excel(self, columns: list = None, sheet_name: str = 'Sheet1', inplace: bool = False) -> None:
        """
        Create an excel document
        :param sheet_name:
        :param inplace: Overwrite if the file exists
        :param columns: If it is blank, a new blank Excel document will be created
        :return:
        """
        _sheet_name = sheet_name
        if os.path.isfile(self.file):
            """
            Overwrite or rebuild files
            """
            if isinstance(inplace, bool) and inplace is False:
                """
                Keep the original file
                """
                raise FileExistenceError(
                    message=f'The {self.file} file already exists, please check whether to replace it, change the '
                            f'parameter inplace=True, or delete this file manually')

        result = pd.DataFrame(columns=columns)
        try:
            result.to_excel(self.file, sheet_name=_sheet_name, engine='openpyxl')
        except Exception:
            raise CreateError

    def add_sheet(self, columns: list = None, sheet_name=None, inplace: bool = False):
        """

        :param columns:
        :param sheet_name:
        :param inplace:
        :return:
        """
        if sheet_name is None:
            sheet_name = []
        # result = pd.DataFrame(columns=columns)
        if os.path.isfile(self.file):
            for title in sheet_name:
                try:
                    self.__wb.create_sheet(
                        title=title,
                        # index=0
                    )
                    # with pd.ExcelWriter(self.file, mode='a', engine='openpyxl') as wf:
                    #     result.to_excel(wf, index=False, header=False, sheet_name=i)
                except Exception:
                    raise CreateError

    def get_sheet_name(self, excel_file: str) -> list:
        if os.path.isfile(excel_file):
            if self.__wb is not None:
                sheet_names = self.__wb.sheetnames
                return sheet_names
        return []

    def remove_sheet(self, sheet_name: str):
        if self.__wb is not None:
            ws = self.__wb[sheet_name]
            self.__wb.remove(ws)
            self.__wb.save(self.file)

    def write_excel(
            self,
            data: list,
            columns: list = None,
            sheet_name: str = 'Sheet1',
            **kwargs):
        """
        Write content to an excel file sheet page
        :param data:
        :param columns:
        :param sheet_name:
        :key mode: w 覆盖式写入内容 ；w+ 追加内容；a 向新的sheet页追加 ;str类型
        :key fill_column: The last element is used as the filling column，单独在最后追加一列；bool
        :return:
        """
        _data = data
        if _data is None:
            raise ParamError(message='Data Parameter exception')
        _file = kwargs.pop('file', self.file)
        if not _file:
            raise ParamError(message=f'Check file path parameters')
        _mode = kwargs.pop('mode', 'w')
        _inplace = kwargs.pop('inplace', False)
        _fill_column = kwargs.pop('fill_column', False)
        _fill_column_data = {}

        if _fill_column is True:
            _fill_column_data = _data.pop()

        for i in _data:
            for k, v in i.items():
                i.update({
                    k: str(v)
                })
        _result = pd.DataFrame(_data, columns=columns)

        for k, v in _fill_column_data.items():
            _result[k] = str(v)

        '''
           If the file does not exist, create the document first
        '''
        self.create_excel(sheet_name=sheet_name, inplace=_inplace)

        try:

            self.__wb = wb = load_workbook(_file)

            '''
               Check whether the sheet name exists
               If it does not exist, the corresponding sheet will be created
            '''
            exits_sheet_name = self.get_sheet_name(excel_file=_file)
            exits_sheet_name = True if sheet_name in exits_sheet_name else False

            if exits_sheet_name is True and _mode == 'w':
                '''
                    First delete the old sheet, then create a new one
                '''
                self.remove_sheet(sheet_name=sheet_name)

            if exits_sheet_name is False and _mode == 'w+':
                self.add_sheet(sheet_name=[sheet_name])

            if exits_sheet_name is True and _mode == 'a':
                raise CreateError(
                    message='When you enable appending in mode a, you must ensure that the sheet does not exist')

            '''
                * When there is no content in the first row, insert the header, and no header will be added for subsequent appending
                Append data to the specified sheet
            '''
            # header_line = self.__wb[sheet_name][1]

            # sheet = wb.active
            sheet = wb.worksheets[wb.sheetnames.index(sheet_name)]

            rows = [row for row in sheet.rows]

            header = kwargs.pop('header', False)
            if header is False:
                exits_header = True
            else:
                if rows:
                    exits_header = True
                else:
                    exits_header = False

            for row in dataframe_to_rows(
                    _result,
                    header=False if exits_header else True,
                    index=False,
            ):
                sheet.append(row)
            wb.save(_file)
            wb.close()
        except Exception:
            raise OpenError
