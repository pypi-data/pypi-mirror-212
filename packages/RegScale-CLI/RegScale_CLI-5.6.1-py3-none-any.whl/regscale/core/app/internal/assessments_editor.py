#!/usr/bin/env python3
# -*- coding: utf-8 -*-
""" Module to allow user to make changes to Assessments in an excel spreadsheet for user friendly experience """

# standard python imports
import json
import os
import shutil
import sys
from pathlib import Path

import click
import pandas as pd
import requests
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Protection, Font, NamedStyle
from openpyxl.worksheet.datavalidation import DataValidation

from regscale.core.app.api import Api
from regscale.core.app.application import Application
from regscale.core.app.logz import create_logger
from regscale.core.app.utils.app_utils import (
    check_file_path,
    error_and_exit,
    reformat_str_date,
)
from regscale.models.app_models.click import regscale_id, regscale_module
from regscale.models.regscale_models.assessment import Assessment
from regscale.models.regscale_models.modules import Modules

logger = create_logger()
app = Application()
config = app.config
api = Api(app)


@click.group(name="assessments")
def assessments():
    """
    [BETA] Performs actions on Assessments Feature
    """


# Make Empty Spreadsheet for creating new assessments.
@assessments.command(name="generate_new_file")
@click.option(
    "--path",
    type=click.Path(exists=False, dir_okay=True, path_type=Path),
    help="Provide the desired path.",
    default=os.path.join(os.getcwd(), "artifacts"),
    required=True,
)
def new_assessment(path: Path):
    """This function will build an excel spreadsheet for users to be able to create new assessments."""
    check_file_path(path)

    # create excel file and setting formatting
    try:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "New_Assessments"

        column_headers = [
            "Title",
            "LeadAssessor",
            "Facility",
            "Organization",
            "AssessmentType",
            "PlannedStart",
            "PlannedFinish",
            "Status",
            "ActualFinish",
            "AssessmentResult",
            "ParentId",
            "ParentModule",
        ]
        for col, val in enumerate(column_headers, start=1):
            worksheet.cell(row=1, column=col).value = val

        for col in ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L"]:
            for cell in worksheet[col]:
                if cell.row == 1:
                    cell.font = Font(bold=True)

        # create and format reference worksheets for dropdowns
        workbook.create_sheet(title="Facilities")
        workbook.create_sheet(title="Organizations")
        workbook.create_sheet(title="Accounts")
        workbook.create_sheet(title="Modules")

        workbook.save(filename=os.path.join(path, "new_assessments.xlsx"))

        # pull in Facility, Organization, Module, and Account Usernames into Excel Spreadsheet to create drop downs
        list_of_modules = Modules().module_names()
        module_names = pd.DataFrame(list_of_modules, columns=["name"])
        with pd.ExcelWriter(
            os.path.join(path, "new_assessments.xlsx"),
            mode="a",
            engine="openpyxl",
            if_sheet_exists="overlay",
        ) as writer:
            get_field_names(field_name="facilities").to_excel(
                writer,
                sheet_name="Facilities",
                index=False,
            )
            get_field_names(field_name="organizations").to_excel(
                writer,
                sheet_name="Organizations",
                index=False,
            )
            get_user_names().to_excel(
                writer,
                sheet_name="Accounts",
                index=False,
            )
            module_names.to_excel(
                writer,
                sheet_name="Modules",
                index=False,
            )

        # Creating data Validation for fields
        workbook = load_workbook(os.path.join(path, "new_assessments.xlsx"))
        worksheet = workbook.active
        facilities_worksheet = workbook["Facilities"]
        accounts_worksheet = workbook["Accounts"]
        organizations_worksheet = workbook["Organizations"]
        modules_worksheet = workbook["Modules"]

        # lock worksheets containing data for dropdowns
        facilities_worksheet.protection.sheet = True
        accounts_worksheet.protection.sheet = True
        organizations_worksheet.protection.sheet = True
        modules_worksheet.protection.sheet = True
        dv1 = DataValidation(
            type="list",
            formula1="=Accounts!$A$2:$A$"
            + str(get_maximum_rows(sheet_object=workbook["Accounts"])),
            allow_blank=False,
            showDropDown=False,
            error="Your entry is not one of the available options",
            errorTitle="Invalid Entry",
            prompt="Please select from the list",
        )
        dv2 = DataValidation(
            type="list",
            formula1="=Facilities!$A$2:$A$"
            + str(get_maximum_rows(sheet_object=workbook["Facilities"])),
            allow_blank=True,
            showDropDown=False,
            error="Your entry is not one of the available options",
            errorTitle="Invalid Entry",
            prompt="Please select from the list",
        )
        dv3 = DataValidation(
            type="list",
            formula1="=Organizations!$A$2:$A$"
            + str(get_maximum_rows(sheet_object=workbook["Organizations"])),
            allow_blank=True,
            showDropDown=False,
            error="Your entry is not one of the available options",
            errorTitle="Invalid Entry",
            prompt="Please select from the list",
        )
        types = '"Control Testing, External Review, Inspection, Internal Audit, Lightning Assessment, Linda\'s metadata for Assessments, Management Assessment, Script/DevOps Check, Walkthrough"'
        dv4 = DataValidation(
            type="list",
            formula1=types,
            allow_blank=False,
            showDropDown=False,
            error="Your entry is not one of the available options",
            errorTitle="Invalid Entry",
            prompt="Please select from the list",
        )
        dv5 = DataValidation(
            type="list",
            formula1='"Scheduled, In Progress, Complete, Cancelled"',
            allow_blank=True,
            showDropDown=False,
            error="Your entry is not one of the available options",
            errorTitle="Invalid Entry",
            prompt="Please select from the list",
        )
        dv6 = DataValidation(
            type="list",
            formula1='"Pass, Fail, N/A, Partial Pass"',
            allow_blank=True,
            showDropDown=False,
            error="Your entry is not one of the available options",
            errorTitle="Invalid Entry",
            prompt="Please select from the list",
        )
        dv7 = DataValidation(
            type="date",
            allow_blank=False,
            showDropDown=False,
            error="Your entry is not a valid option",
            errorTitle="Invalid Entry",
            prompt="Please enter valid date mm/dd/yyyy",
        )
        dv8 = DataValidation(
            type="list",
            formula1="=Modules!$A$2:$A$"
            + str(get_maximum_rows(sheet_object=workbook["Modules"])),
            allow_blank=True,
            showDropDown=False,
            error="Your entry is not a valid option",
            errorTitle="Invalid Entry",
            prompt="Please select from the list",
        )
        dv9 = DataValidation(
            type="date",
            allow_blank=True,
            showDropDown=False,
            error="Your entry is not a valid option",
            errorTitle="Invalid Entry",
            prompt="Please enter valid date mm/dd/yyyy",
        )
        worksheet.add_data_validation(dv1)
        worksheet.add_data_validation(dv2)
        worksheet.add_data_validation(dv3)
        worksheet.add_data_validation(dv4)
        worksheet.add_data_validation(dv5)
        worksheet.add_data_validation(dv6)
        worksheet.add_data_validation(dv7)
        worksheet.add_data_validation(dv8)
        worksheet.add_data_validation(dv9)
        dv1.add("B2:B1048576")
        dv2.add("C2:C1048576")
        dv3.add("D2:D1048576")
        dv4.add("E2:E1048576")
        dv5.add("H2:H1048576")
        dv6.add("J2:J1048576")
        dv7.add("F2:F1048576")
        dv7.add("G2:G1048576")
        dv9.add("I2:I1048576")
        dv8.add("L2:L1048576")

        workbook.save(filename=os.path.join(path, "new_assessments.xlsx"))

        # Freezing top row and adding data style to date columns to assure validation

        workbook = load_workbook(os.path.join(path, "new_assessments.xlsx"))
        worksheet = workbook.active
        freeze_range = worksheet.cell(2, 14)
        worksheet.freeze_panes = freeze_range
        date_style = NamedStyle(name="date_style", number_format="mm/dd/yyyy")
        workbook.add_named_style(date_style)

        for col in ["F", "G", "I"]:  # Columns to edit
            for cell in worksheet[col]:
                if cell.row > 1:
                    cell.style = date_style

        # Adjusting width of columns

        for col in worksheet.columns:
            max_length = 0
            column = col[0].column_letter  # Get the column name
            for cell in col:
                try:  # Necessary to avoid error on empty cells
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except OSError:
                    logger.error("Cell adjustment failed due to empty cells.")
            adjusted_width = (max_length + 2) * 1.2
            worksheet.column_dimensions[column].width = adjusted_width

        workbook.save(filename=os.path.join(path, "new_assessments.xlsx"))

    except OSError:
        logger.error(f"Creation of the directory {path} failed.")

    return logger.info(
        "Your excel workbook has been created. Please open the new_assessments workbook and add new assessments."
    )


@assessments.command(name="generate")
@regscale_id()
@regscale_module()
@click.option(
    "--path",
    type=click.Path(exists=False, dir_okay=True, path_type=Path),
    help="Provide the desired path.",
    default=os.path.join(os.getcwd(), "artifacts"),
    required=True,
)
def all_assessments(regscale_id: str, regscale_module: str, path: Path):
    """This function will build and populate a spreadsheet of all assessments
    with the selected RegScale Parent Id and RegScale Module for users to any necessary edits.
    """
    try:
        body = """
                query {
                      assessments (skip: 0, take: 50, where: {parentId: {eq: parent_id} parentModule: {eq: "parent_module"}}) {
                        items {
                          id
                          title
                          leadAssessor {
                            firstName
                            lastName
                            userName
                          }
                          facility {
                            name
                          }
                          org {
                            name
                          }
                          assessmentType
                          plannedStart
                          plannedFinish
                          status
                          actualFinish
                          assessmentResult
                          parentId
                          parentModule
                        }
                        totalCount
                        pageInfo {
                          hasNextPage
                        }
                    }
                }
                    """.replace(
            "parent_module", regscale_module
        ).replace(
            "parent_id", str(regscale_id)
        )
        existing_assessment_data = api.graph(query=body)

    except requests.RequestException as ex:
        error_and_exit("Unable to retrieve assessment list from RegScale.\n %s", ex)

    if (
        existing_assessment_data["assessments"]["totalCount"] > 0
    ):  # Checking to see if assessment exists for selected RegScale Id and RegScale Module.
        check_file_path(path)

        # Loading data from db into two workbooks.
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = f"Assessments({regscale_id}_{regscale_module})"
        workbook.create_sheet(title="Facilities")
        workbook.create_sheet(title="Organizations")
        workbook.create_sheet(title="Accounts")
        workbook.save(filename=os.path.join(path, "all_assessments.xlsx"))
        shutil.copy(
            os.path.join(path, "all_assessments.xlsx"),
            os.path.join(path, "old_assessments.xlsx"),
        )

        raw_data = existing_assessment_data["assessments"]["items"]
        assessments_data = []
        for a in raw_data:
            Id = a["id"]
            Title = a["title"]
            LeadAssessor = (
                str(a["leadAssessor"]["lastName"])
                + ", "
                + str(a["leadAssessor"]["firstName"])
                + " ("
                + str(a["leadAssessor"]["userName"])
                + ")"
            )
            Facility = a["facility"]["name"] if a["facility"] else "None"
            Organization = a["org"]["name"] if a["org"] else "None"
            AssessmentType = a["assessmentType"]
            PlannedStart = reformat_str_date(a["plannedStart"])
            PlannedFinish = reformat_str_date(a["plannedFinish"])
            Status = a["status"]
            ActualFinish = (
                reformat_str_date(a["actualFinish"]) if a["actualFinish"] else "None"
            )
            AssessmentResult = a["assessmentResult"] or "None"
            ParentId = a["parentId"]
            ParentModule = a["parentModule"]

            assessments_data.append(
                [
                    Id,
                    Title,
                    LeadAssessor,
                    Facility,
                    Organization,
                    AssessmentType,
                    PlannedStart,
                    PlannedFinish,
                    Status,
                    ActualFinish,
                    AssessmentResult,
                    ParentId,
                    ParentModule,
                ]
            )

        all_ass_df = pd.DataFrame(
            assessments_data,
            columns=[
                "Id",
                "Title",
                "LeadAssessor",
                "Facility",
                "Organization",
                "AssessmentType",
                "PlannedStart",
                "PlannedFinish",
                "Status",
                "ActualFinish",
                "AssessmentResult",
                "ParentId",
                "ParentModule",
            ],
        )

        with pd.ExcelWriter(
            os.path.join(path, "all_assessments.xlsx"),
            mode="a",
            engine="openpyxl",
            if_sheet_exists="overlay",
        ) as writer:
            all_ass_df.to_excel(
                writer,
                sheet_name=f"Assessments({regscale_id}_{regscale_module})",
                index=False,
            )
        with pd.ExcelWriter(
            os.path.join(path, "old_assessments.xlsx"),
            mode="a",
            engine="openpyxl",
            if_sheet_exists="overlay",
        ) as writer:
            all_ass_df.to_excel(
                writer,
                sheet_name=f"Assessments{regscale_id}_{regscale_module})",
                index=False,
            )

        # Pulling in Facility Names into Excel Spreadsheet to create dropdown.
        with pd.ExcelWriter(
            os.path.join(path, "all_assessments.xlsx"),
            mode="a",
            engine="openpyxl",
            if_sheet_exists="overlay",
        ) as writer:
            get_field_names(field_name="facilities").to_excel(
                writer,
                sheet_name="Facilities",
                index=False,
            )
            get_field_names(field_name="organizations").to_excel(
                writer,
                sheet_name="Organizations",
                index=False,
            )
            get_user_names().to_excel(
                writer,
                sheet_name="Accounts",
                index=False,
            )

        # Adding protection to "old_assessments.xlsx" file that will be used as reference.
        workbook2 = load_workbook(os.path.join(path, "old_assessments.xlsx"))
        worksheet2 = workbook2.active
        worksheet2.protection.sheet = True
        workbook2.save(filename=os.path.join(path, "old_assessments.xlsx"))

        # Adding Data Validation to "all_assessments.xlsx" file to be adjusted internally.
        workbook = load_workbook(os.path.join(path, "all_assessments.xlsx"))
        worksheet = workbook.active
        facilities_worksheet = workbook["Facilities"]
        accounts_worksheet = workbook["Accounts"]
        organizations_worksheet = workbook["Organizations"]

        # lock worksheets containing data for dropdowns
        facilities_worksheet.protection.sheet = True
        accounts_worksheet.protection.sheet = True
        organizations_worksheet.protection.sheet = True
        worksheet.protection.sheet = True

        # Unlocking cells that can be edited in each Assessment.
        for col in ["C", "D", "E", "F", "G", "H", "I", "J", "K"]:  # Columns to edit
            for cell in worksheet[col]:
                cell.protection = Protection(locked=False)

        dv1 = DataValidation(
            type="list",
            formula1="=Accounts!$A$2:$A$"
            + str(get_maximum_rows(sheet_object=workbook["Accounts"])),
            allow_blank=False,
            showDropDown=False,
            error="Your entry is not one of the available options",
            errorTitle="Invalid Entry",
            prompt="Please select from the list",
        )
        dv2 = DataValidation(
            type="list",
            formula1="=Facilities!$A$2:$A$"
            + str(get_maximum_rows(sheet_object=workbook["Facilities"])),
            allow_blank=True,
            showDropDown=False,
            error="Your entry is not one of the available options",
            errorTitle="Invalid Entry",
            prompt="Please select from the list",
        )
        dv3 = DataValidation(
            type="list",
            formula1="=Organizations!$A$2:$A$"
            + str(get_maximum_rows(sheet_object=workbook["Organizations"])),
            allow_blank=True,
            showDropDown=False,
            error="Your entry is not one of the available options",
            errorTitle="Invalid Entry",
            prompt="Please select from the list",
        )
        types = '"Control Testing, External Review, Inspection, Internal Audit, Lightning Assessment, Linda\'s metadata for Assessments, Management Assessment, Script/DevOps Check, Walkthrough"'
        dv4 = DataValidation(
            type="list",
            formula1=types,
            allow_blank=False,
            showDropDown=False,
            error="Your entry is not one of the available options",
            errorTitle="Invalid Entry",
            prompt="Please select from the list",
        )
        dv5 = DataValidation(
            type="list",
            formula1='"Scheduled, In Progress, Complete, Cancelled"',
            allow_blank=True,
            showDropDown=False,
            error="Your entry is not one of the available options",
            errorTitle="Invalid Entry",
            prompt="Please select from the list",
        )
        dv6 = DataValidation(
            type="list",
            formula1='"Pass, Fail, N/A, Partial Pass"',
            allow_blank=True,
            showDropDown=False,
            error="Your entry is not one of the available options",
            errorTitle="Invalid Entry",
            prompt="Please select from the list",
        )
        dv7 = DataValidation(
            type="date",
            allow_blank=False,
            showDropDown=False,
            error="Your entry is not a valid option",
            errorTitle="Invalid Entry",
            prompt="Please enter valid date mm/dd/yyyy",
        )
        dv8 = DataValidation(
            type="date",
            allow_blank=True,
            showDropDown=False,
            error="Your entry is not a valid option",
            errorTitle="Invalid Entry",
            prompt="Please enter valid date mm/dd/yyyy",
        )
        worksheet.add_data_validation(dv1)
        worksheet.add_data_validation(dv2)
        worksheet.add_data_validation(dv3)
        worksheet.add_data_validation(dv4)
        worksheet.add_data_validation(dv5)
        worksheet.add_data_validation(dv6)
        worksheet.add_data_validation(dv7)
        worksheet.add_data_validation(dv8)
        dv1.add("C2:C1048576")
        dv2.add("D2:D1048576")
        dv3.add("E2:E1048576")
        dv4.add("F2:F1048576")
        dv5.add("I2:I1048576")
        dv6.add("K2:K1048576")
        dv7.add("G2:G1048576")
        dv7.add("H2:H1048576")
        dv8.add("J2:J1048576")

        for col in worksheet.columns:
            max_length = 0
            column = col[0].column_letter  # Get the column name
            for cell in col:
                try:  # Necessary to avoid error on empty cells
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except OSError:
                    logger.error("Cell adjustment failed due to empty cells.")
            adjusted_width = (max_length + 2) * 1.2
            worksheet.column_dimensions[column].width = adjusted_width

        # Worksheet freeze top row
        freeze_range = worksheet.cell(2, 16)
        worksheet.freeze_panes = freeze_range
        date_style = NamedStyle(name="date_style", number_format="mm/dd/yyyy")
        workbook.add_named_style(date_style)

        # Adding Date Style to Worksheet
        for col in ["G", "H", "J"]:  # Columns to edit
            for cell in worksheet[col]:
                if cell.row > 1:
                    cell.style = date_style

        workbook.save(filename=os.path.join(path, "all_assessments.xlsx"))

    else:
        logger.info(
            "Please check your selections for RegScale Id and RegScale Module and try again."
        )
        error_and_exit(
            "There was an error creating your workbook for the given RegScale Id and RegScale Module."
        )

    return logger.info(
        "Your data has been loaded into your excel workbook. Please open the all_assessments workbook and make your desired changes."
    )


@assessments.command(name="load")
@click.option(
    "--path",
    type=click.Path(exists=False, dir_okay=True, path_type=Path),
    help="Provide the desired path.",
    default=os.path.join(os.getcwd(), "artifacts"),
    required=True,
)
def upload_data(path: Path) -> None:
    """
    This function uploads updated assessments and new assessments to the RegScale database from the Excel files that users have edited.
    """
    # Checking if new assessments have been created and updating RegScale database.
    if os.path.isfile(os.path.join(path, "new_assessments.xlsx")):
        new_files = os.path.join(path, "new_assessments.xlsx")
        new = pd.read_excel(new_files)
        new["Facility"] = new["Facility"].fillna("None")
        new["Organization"] = new["Organization"].fillna("None")
        new["ParentId"] = new["ParentId"].fillna("None")
        new["ParentModule"] = new["ParentModule"].fillna("None")
        new["AssessmentResult"] = new["AssessmentResult"].fillna("N/A")
        facilities = pd.read_excel(new_files, sheet_name="Facilities")
        facilities = facilities.rename(columns={"name": "Facility", "id": "FacilityId"})
        organizations = pd.read_excel(new_files, sheet_name="Organizations")
        organizations = organizations.rename(
            columns={"name": "Organization", "id": "OrganizationId"}
        )
        accounts = pd.read_excel(new_files, sheet_name="Accounts")
        new = new.merge(accounts, how="left", on="LeadAssessor")
        new = new.merge(facilities, how="left", on="Facility")
        new = new.merge(organizations, how="left", on="Organization")
        new = new.T.to_dict()
        new_assessments = [
            Assessment(
                leadAssessorId=value["LeadAssessorId"],
                title=value["Title"],
                assessmentType=value["AssessmentType"],
                plannedStart=value["PlannedStart"],
                plannedFinish=value["PlannedFinish"],
                status=value["Status"],
                facilityId=value["FacilityId"],
                orgId=value["OrganizationId"],
                assessmentResult=value["AssessmentResult"],
                actualFinish=value["ActualFinish"],
                parentId=value["ParentId"],
                parentModule=value["ParentModule"],
            )
            for value in new.values()
        ]
        new_load = json.dumps(new_assessments)
        try:
            api.post(
                url=config["domain"] + "/api/assessments/batchCreate",
                json=new_load,
            )
            logger.info(
                "%s total assessments were added to RegScale database.",
                str(len(new_load)),
            )
        except requests.exceptions.RequestException as ex:
            logger.error(
                f"Unable to add {len(new_load)} assessment(s) to RegScale.",
                ex,
            )
    elif os.path.isfile(os.path.join(path, "all_assessments")):
        # Checking all_assessments file for differences before updating database

        df1 = load_workbook(os.path.join(path, "old_assessments.xlsx"))
        ws1 = df1.active
        data1 = ws1.values
        columns1 = next(data1)[:]
        df1 = pd.DataFrame(data1, columns=columns1)

        df2 = load_workbook(os.path.join(path, "all_assessments.xlsx"))
        ws2 = df2.active
        data2 = ws2.values
        columns2 = next(data2)[:]
        df2 = pd.DataFrame(data2, columns=columns2)

        if df1.equals(df2):
            logger.info("No differences detected.")
            sys.exit(1)
        else:
            logger.warning("Differences found!")

        updated_files = os.path.join(path, "all_assessments.xlsx")
        updated = pd.read_excel(updated_files)
        updated["Facility"] = updated["Facility"].fillna("None")
        updated["Organization"] = updated["Organization"].fillna("None")
        updated["AssessmentResult"] = updated["AssessmentResult"].fillna("N/A")
        facilities = pd.read_excel(updated_files, sheet_name="Facilities")
        facilities = facilities.rename(columns={"name": "Facility", "id": "FacilityId"})
        organizations = pd.read_excel(updated_files, sheet_name="Organizations")
        organizations = organizations.rename(
            columns={"name": "Organization", "id": "OrganizationId"}
        )
        accounts = pd.read_excel(updated_files, sheet_name="Accounts")
        updated = updated.merge(accounts, how="left", on="LeadAssessor")
        updated = updated.merge(facilities, how="left", on="Facility")
        updated = updated.merge(organizations, how="left", on="Organization")
        updated = updated.T.to_dict()
        updated_assessments = []
        for value in updated.values():
            updated_assessments.append(
                Assessment(
                    leadAssessorId=value["LeadAssessorId"],
                    id=value["Id"],
                    title=value["Title"],
                    assessmentType=value["AssessmentType"],
                    plannedStart=value["PlannedStart"],
                    plannedFinish=value["PlannedFinish"],
                    status=value["Status"],
                    facilityId=value["FacilityId"],
                    orgId=value["OrganizationId"],
                    assessmentResult=value["AssessmentResult"],
                    actualFinish=value["ActualFinish"],
                    parentId=value["ParentId"],
                    parentModule=value["ParentModule"],
                )
            )

            updated_load = json.dumps(updated_assessments)
            try:
                response = api.post(
                    url=config["domain"] + f"/api/assessments/{id}",
                    json=updated_load,
                )
                if response.ok:
                    logger.info(f"Successfully updated assessment #{id} in RegScale.")
                else:
                    logger.warning(f"Unable to update assessment #{id} in RegScale.")
            except requests.exceptions.RequestException as ex:
                logger.error(
                    "Unable to update assessments in RegScale database. \n %s", ex
                )
    os.remove(os.path.join(path, "all_assessments.xlsx"))
    os.remove(os.path.join(path, "old_assessments.xlsx"))
    os.remove(os.path.join(path, "new_assessments.xlsx"))
    return logger.info(
        "Assessment files have been deleted. Changes made to existing files can be seen in differences.txt file. Thank you!"
    )


def get_maximum_rows(*, sheet_object):
    """This function finds the last row containing data in a spreadsheet
    :param sheet_object: excel worksheet to be referenced
    :return: integer representing last row with data in spreadsheet
    :rtype: int
    """
    return sum(
        any(col.value is not None for col in row)
        for max_row, row in enumerate(sheet_object, 1)
    )


def get_field_names(field_name):
    """
    This function uses GraphQL to retrieve all names of a given parent table in database
    :return: pandas dataframe with facility names
    :rtype: pd.dataframe
    """
    body = """
    query {
        field_name(skip: 0, take: 50, order: {name: ASC}, ) {
            items {
                name
                id
            }
            totalCount
            pageInfo {
                hasNextPage
            }
        }
    }
    """.replace(
        "field_name", field_name
    )

    field_items = api.graph(query=body)
    names = field_items[str(field_name)]["items"]
    field_names = [[i["name"], i["id"]] for i in names]
    all_names = pd.DataFrame(field_names, index=None, columns=["name", "id"])
    all_names.loc[len(all_names.index)] = ["None", "None"]

    return all_names


def get_user_names():
    """This function uses API Endpoint to retrieve all user names in database
    :return: pandas dataframe with usernames
    :rtype: pd.dataframe
    """
    accounts = api.get(url=config["domain"] + "/api/accounts").json()

    user_names = [[item["name"], item["id"]] for item in accounts]
    return pd.DataFrame(
        user_names,
        index=None,
        columns=["LeadAssessor", "LeadAssessorId"],
    )
