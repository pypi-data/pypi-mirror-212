#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Module to allow user to make changes to Control Implementations in an Excel spreadsheet for a user-friendly experience
"""

# standard python imports
import json
import os
import shutil
import sys
from pathlib import Path

import click
import numpy as np
import pandas as pd
import requests
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Protection
from openpyxl.worksheet.datavalidation import DataValidation

from regscale.core.app.api import Api
from regscale.core.app.application import Application
from regscale.core.app.logz import create_logger
from regscale.core.app.utils.app_utils import error_and_exit
from regscale.models.app_models.click import regscale_id, regscale_module
from regscale.models.regscale_models.control_implementation import (
    ControlImplementation,
    Control,
)

logger = create_logger()
app = Application()
config = app.config
api = Api(app)


@click.group(name="control_editor")
def control_editor():
    """
    Performs actions on Control Editor Feature!
    """


# Get data and pull into Excel worksheets.


@control_editor.command(name="data_download")
@regscale_id()
@regscale_module()
@click.option(
    "--path",
    type=click.Path(exists=False, dir_okay=True, path_type=Path),
    help="Provide the desired path.",
    default=os.path.join(os.getcwd(), "artifacts"),
    required=True,
)
def data_load(regscale_id: str, regscale_module: str, path: Path):
    """
    This function will build and populate a spreadsheet of all control implementations
    with the selected RegScale Parent Id and RegScale Module.
    """

    # Making directory for files

    if os.path.isdir(path) is False:
        try:
            os.mkdir(path)
            workbook = Workbook()
            ws = workbook.active
            ws.title = f"Impls_PId({regscale_id}_{regscale_module})"
            workbook.save(filename=os.path.join(path, "all_implementations.xlsx"))
            shutil.copy(
                os.path.join(path, "all_implementations.xlsx"),
                os.path.join(path, "old_implementations.xlsx"),
            )

            # Loading data from RegScale database into two workbooks.

            try:
                body = """
                        query{
                            controlImplementations (skip: 0, take: 50, where: {parentId: {eq: parent_id} parentModule: {eq: "parent_module"}}) {
                                items {
                                id
                                controlID
                                controlOwnerId
                                control {
                                    title
                                    description
                                }
                                status
                                policy
                                implementation
                                responsibility
                                inheritable
                                parentId
                                parentModule
                                }
                                totalCount
                                pageInfo {
                                    hasNextPage
                                }
                            }
                            }""".replace(
                    "parent_module", regscale_module
                ).replace(
                    "parent_id", regscale_id
                )

                existing_implementation_data = api.graph(query=body)

            except requests.RequestException as ex:
                error_and_exit(
                    "Unable to retrieve assessment list from RegScale.\n %s", ex
                )

            if existing_implementation_data["controlImplementations"]["totalCount"] > 0:
                raw_data = existing_implementation_data["controlImplementations"][
                    "items"
                ]

                all_imps = []
                for item in raw_data:
                    Id = item["id"]
                    ControlId = item["controlID"]
                    ControlOwnerId = item["controlOwnerId"]
                    ControlName = item["control"]["title"]
                    Description = item["control"]["description"]
                    Status = item["status"]
                    Policy = item["policy"]
                    Implementation = item["implementation"]
                    Responsibility = item["responsibility"]
                    Inheritable = item["inheritable"]

                    all_imps.append(
                        [
                            Id,
                            ControlId,
                            ControlOwnerId,
                            ControlName,
                            Description,
                            Status,
                            Policy,
                            Implementation,
                            Responsibility,
                            Inheritable,
                        ]
                    )

                all_imps_df = pd.DataFrame(
                    all_imps,
                    columns=[
                        "Id",
                        "ControlId",
                        "ControlOwnerId",
                        "ControlName",
                        "Description",
                        "Status",
                        "Policy",
                        "Implementation",
                        "Responsibility",
                        "Inheritable",
                    ],
                )

                with pd.ExcelWriter(
                    os.path.join(path, "all_implementations.xlsx"),
                    mode="w",
                    engine="openpyxl",
                ) as writer:
                    all_imps_df.to_excel(
                        writer,
                        sheet_name=f"Impls_PId({regscale_id}_{regscale_module})",
                        index=False,
                    )

                with pd.ExcelWriter(
                    os.path.join(path, "old_implementations.xlsx"),
                    mode="w",
                    engine="openpyxl",
                ) as writer:
                    all_imps_df.to_excel(
                        writer,
                        sheet_name=f"Impls_PId({regscale_id}_{regscale_module})",
                        index=False,
                    )
            else:
                logger.info(
                    "Please check your selections for RegScale Id and RegScale Module and try again."
                )
                error_and_exit(
                    "There was an error creating your workbook for the given RegScale Id and RegScale Module."
                )

            # Adding Data validation to "old_implementations.xlsx" file that will be used as reference.

            workbook2 = load_workbook(os.path.join(path, "old_implementations.xlsx"))
            worksheet2 = workbook2.active
            worksheet2.protection.sheet = True
            workbook2.save(filename=os.path.join(path, "old_implementations.xlsx"))

            # Adding Data Validation to "all_implementations.xlsx" file to be adjusted internally by clients.

            workbook = load_workbook(os.path.join(path, "all_implementations.xlsx"))
            worksheet = workbook.active
            worksheet.protection.sheet = True

            dv1 = DataValidation(
                type="list",
                formula1='"Not Implemented, Fully Implemented, In Remediation, Not Applicable, Inherited, Planned"',
                allow_blank=True,
                showDropDown=False,
                error="Your entry is not one of the available options",
                errorTitle="Invalid Entry",
                prompt="Please select from the list",
            )
            dv2 = DataValidation(
                type="list",
                formula1='"Provider, Customer, Shared, Not Applicable"',
                allow_blank=True,
                showDropDown=False,
                error="Your entry is not one of the available options",
                errorTitle="Invalid Entry",
                prompt="Please select from the list",
            )
            dv3 = DataValidation(
                type="list", formula1='"TRUE, FALSE"', allow_blank=True
            )

            worksheet.add_data_validation(dv1)
            worksheet.add_data_validation(dv2)
            worksheet.add_data_validation(dv3)
            dv1.add("F2:F1048576")
            dv2.add("I2:I1048576")
            dv3.add("J2:J1048576")

            for col in worksheet.columns:
                max_length = 0
                column = col[0].column_letter  # Get the column name
                for cell in col:
                    try:  # Necessary to avoid error on empty cells
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except OSError:
                        logger.error("Cell adjustment failed due to empty cells.")
                adjusted_width = (max_length + 2) * 1.2
                worksheet.column_dimensions[column].width = adjusted_width

            for col in ["F", "G", "H", "I", "J"]:
                for cell in worksheet[col]:
                    cell.protection = Protection(locked=False)

            workbook.save(filename=os.path.join(path, "all_implementations.xlsx"))

        except OSError:
            logger.error(f"Creation of the directory {path} failed.")
    else:
        logger.info(f"Successfully created the directory {path}.")
        logger.info("All files are located within directory.")
        sys.exit()

    return logger.info(
        "Your data has been loaded into your excel workbook. Please open the all_implementations workbook and make your desired changes."
    )


# Save Spreadsheet if file changed, append Update API changes that were manually entered in an Excel worksheet


@control_editor.command(name="data_upload")
@click.option(
    "--path",
    type=click.Path(exists=False, dir_okay=True, path_type=Path),
    help="Provide the desired path.",
    default=os.path.join(os.getcwd(), "artifacts"),
    required=True,
)
@click.option(
    "--skip_prompt",
    type=click.BOOL,
    help="To Skip (Y/N) Prompt, input True.",
    default=False,
    required=False,
)
def db_update(path: Path, skip_prompt: bool):
    """
    This function will check changes made to spreadsheet and upload any changes made to RegScale database.

    """
    logger.info(
        "Proceed only after you have made the necessary changes and have saved file."
    )

    x = "y" if skip_prompt else input("Ready to Proceed (Y/N): ").lower()

    if x[0] == "y":
        df1 = load_workbook(os.path.join(path, "all_implementations.xlsx"))
        ws1 = df1.active

        sheet_name = df1.sheetnames[0]
        sheet_name = sheet_name[sheet_name.find("(") + 1 : sheet_name.find(")")].split(
            "_"
        )
        regscale_parent_id, regscale_module = set(sheet_name)

        data1 = ws1.values
        columns1 = next(data1)[0:]
        df1 = pd.DataFrame(data1, columns=columns1)

        df2 = load_workbook(os.path.join(path, "old_implementations.xlsx"))
        ws2 = df2.active
        data2 = ws2.values
        columns2 = next(data2)[0:]
        df2 = pd.DataFrame(data2, columns=columns2)

        data_frame_same = df1.equals(df2)  # Files being compared here.

        with open(
            os.path.join(path, "differences.txt"),
            "w+",
        ) as f:
            f.truncate(0)

        if data_frame_same:
            logger.info("No differences detected.")

        else:
            upload_data(regscale_parent_id, regscale_module, path)
            logger.warning("*** WARNING *** Differences Found.")

            # Logs changes to txt file

            diff_mask = (df1 != df2) & ~(df1.isnull() & df2.isnull())
            ne_stacked = diff_mask.stack()
            changed = ne_stacked[ne_stacked]
            changed.index.names = ["row", "col"]
            difference_locations = np.where(diff_mask)
            changed_from = df1.values[difference_locations]
            changed_to = df2.values[difference_locations]
            changes = pd.DataFrame(
                {"from": changed_from, "to": changed_to}, index=changed.index
            )
            changes.to_csv(
                os.path.join(path, "differences.txt"),
                header=None,
                index=None,
                sep=" ",
                mode="a",
            )
    return logger.info(
        "Please check differences.txt file located in artifacts folder of current working directory to see changes made."
    )


def upload_data(regscale_parent_id: int, regscale_module: str, path: Path) -> None:
    """
    Batch uploads updated control implementation statements to the provided RegScale parent ID
    :param int regscale_parent_id: RegScale parent ID
    :param str regscale_module: RegScale parent module
    :param Path path: file path where control spreadsheet resides
    :raises: requests.exceptions.RequestException if API call encountered an error
    :return: None
    """
    updated_implementations = []
    reader = pd.read_excel(os.path.join(path, "all_implementations.xlsx"))
    updates = reader.T.to_dict()
    for i in updates.values():
        updated_implementations.append(
            ControlImplementation(
                id=i["Id"],
                controlOwnerId=i["ControlOwnerId"],
                control=Control(
                    title=i["ControlName"], description=i["Description"]
                ).dict(),
                status=i["Status"],
                implementation=i["Implementation"],
                policy=i["Policy"],
                controlID=i["ControlId"],
                responsibility=i["Responsibility"],
                parentId=int(regscale_parent_id),
                parentModule=regscale_module,
                inheritable=i["Inheritable"],
            )
        )
    new_implementations = json.dumps(updated_implementations)
    try:
        api.post(
            url=config["domain"] + "/api/controlImplementation/batchUpdate",
            json=new_implementations,
        )
        logger.info(
            "%s total %s for Parent ID: %s in RegScale were updated.",
            len(new_implementations),
            regscale_module,
            regscale_parent_id,
        )
    except requests.exceptions.RequestException as ex:
        logger.error(
            "Unable to update %s for ParentId: %s in RegScale \n %s",
            regscale_module,
            regscale_parent_id,
            ex,
        )


# Delete and remove files from user's system.


@control_editor.command(name="delete_files")
@click.option(
    "--path",
    type=click.Path(exists=False, dir_okay=True, path_type=Path),
    help="Provide the desired path.",
    default=os.path.join(os.getcwd(), "artifacts"),
    required=True,
)
def delete_file(path: Path):
    """
    Deletes files used during the process.

    """
    os.remove(os.path.join(path, "all_implementations.xlsx"))
    os.remove(os.path.join(path, "old_implementations.xlsx"))
    os.remove(os.path.join(path, "differences.txt"))
    os.rmdir(path)
    return logger.info("Files have been deleted. Thank you.")
