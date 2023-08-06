#!/usr/bin/env python3
# -*- coding: utf-8 -*-
""" Dataclass for a control_editor.py module """

import json
import os
import shutil
import sys
from pathlib import Path

import numpy as np
import openpyxl
import pandas as pd
import requests
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Protection
from openpyxl.worksheet.datavalidation import DataValidation

from regscale.core.app.api import Api
from regscale.core.app.application import Application
from regscale.core.app.internal.assessments_editor import upload_data
from regscale.core.app import create_logger

app = Application()
config = app.config
api = Api(app)


class Control_Editor:
    def __init__(self, file: str, app: Application):
        self.file = file
        self.logger = create_logger()
        self.app = app

    def data_load(
        self, regscale_parent_id: int, regscale_module: str, path: Path
    ) -> None:
        """
        This function will build and populate a spreadsheet of all control implementations
        with the selected RegScale and RegScale Module
        :param int regscale_parent_id: RegScale parent ID
        :param str regscale_module: RegScale Module
        :param path: file path where control spreadsheet resides
        :return: None
        """
        if os.path.isdir(path) is False:
            try:
                os.mkdir(path)
                workbook = Workbook()
                ws = workbook.active
                ws.title = f"Impls_PId({regscale_parent_id}_{regscale_module})"
                workbook.save(filename=os.path.join(path, "all_implementations.xlsx"))
                shutil.copy(
                    os.path.join(path, "all_implementations.xlsx"),
                    os.path.join(path, "old_implementations.xlsx"),
                )

                # Loading data from db in to two workbooks.

                imp_r = api.get(
                    url=config["domain"]
                    + f"/api/controlImplementation/getAllByParent/{regscale_parent_id}/{regscale_module}"
                ).json()

                all_imps = []
                for item in imp_r:
                    Id = item["id"]
                    ControlId = item["controlID"]
                    ControlName = item["controlName"]
                    Description = item["controlTitle"]
                    Status = item["status"]
                    Policy = item["policy"]
                    Implementation = item["implementation"]
                    Responsibility = item["responsibility"]
                    Inheritable = item["inheritable"]

                    all_imps.append(
                        [
                            Id,
                            ControlId,
                            ControlName,
                            Description,
                            Status,
                            Policy,
                            Implementation,
                            Responsibility,
                            Inheritable,
                        ]
                    )

                all_imps_df = pd.DataFrame(
                    all_imps,
                    columns=[
                        "Id",
                        "ControlId",
                        "ControlName",
                        "Description",
                        "Status",
                        "Policy",
                        "Implementation",
                        "Responsibility",
                        "Inheritable",
                    ],
                )

                with pd.ExcelWriter(
                    os.path.join(path, "all_implementations.xlsx"),
                    mode="w",
                    engine="openpyxl",
                ) as writer:
                    all_imps_df.to_excel(
                        writer,
                        sheet_name=f"Impls_PId({regscale_parent_id}_{regscale_module})",
                        index=False,
                    )

                with pd.ExcelWriter(
                    os.path.join(path, "old_implementations.xlsx"),
                    mode="w",
                    engine="openpyxl",
                ) as writer:
                    all_imps_df.to_excel(
                        writer,
                        sheet_name=f"Impls_PId({regscale_parent_id}_{regscale_module})",
                        index=False,
                    )

                # Adding Data validation to "old_implementations.xlsx" file that will be used as reference.

                workbook2 = load_workbook(
                    os.path.join(path, "old_implementations.xlsx")
                )
                worksheet2 = workbook2.active
                worksheet2.protection.sheet = True
                workbook2.save(filename=os.path.join(path, "old_implementations.xlsx"))

                # Adding Data Validation to "all_implementations.xlsx" file to be adjusted internally.

                workbook = load_workbook(os.path.join(path, "all_implementations.xlsx"))
                worksheet = workbook.active
                worksheet.protection.sheet = True

                dv1 = DataValidation(
                    type="list",
                    formula1='"Not Implemented, Fully Implemented, In Remediation, Not Applicable, Inherited, Planned"',
                    allow_blank=True,
                    showDropDown=False,
                    error="Your entry is not one of the available options",
                    errorTitle="Invalid Entry",
                    prompt="Please select from the list",
                )
                dv2 = DataValidation(
                    type="list",
                    formula1='"Provider, Customer, Shared, Not Applicable"',
                    allow_blank=True,
                    showDropDown=False,
                    error="Your entry is not one of the available options",
                    errorTitle="Invalid Entry",
                    prompt="Please select from the list",
                )
                dv3 = DataValidation(
                    type="list", formula1='"TRUE, FALSE"', allow_blank=True
                )

                worksheet.add_data_validation(dv1)
                worksheet.add_data_validation(dv2)
                worksheet.add_data_validation(dv3)
                dv1.add("E2:E1048576")
                dv2.add("H2:H1048576")
                dv3.add("I2:I1048576")

                for col in worksheet.columns:
                    max_length = 0
                    column = col[0].column_letter  # Get the column name
                    for cell in col:
                        try:  # Necessary to avoid error on empty cells
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except OSError:
                            self.logger.error(
                                "Cell adjustment failed due to empty cells."
                            )
                    adjusted_width = (max_length + 2) * 1.2
                    worksheet.column_dimensions[column].width = adjusted_width

                for col in ["E", "F", "G", "H", "I"]:
                    for cell in worksheet[col]:
                        cell.protection = Protection(locked=False)

                workbook.save(filename=os.path.join(path, "all_implementations.xlsx"))

            except OSError:
                self.logger.error("Creation of the directory %s failed." % path)
            else:
                self.logger.info("Successfully created the directory %s." % path)
                self.logger.info("All files are located within directory.")
                sys.exit()

        return self.logger(
            "Your data has been loaded into your excel workbook. Please open the all_implementations workbook and make your desired changes."
        )

    def db_update(
        self, regscale_parent_id: int, regscale_module: str, path: Path
    ) -> None:
        """
        This function will check changes made to spreadsheet and upload any changes made to RegScale database
        :param int regscale_parent_id: ID of parent item in RegScale
        :param str regscale_module: Parent module from RegScale
        :param path: file path where control spreadsheet resides
        :return: None
        """
        self.logger.info(
            "Proceed only after you have made the necessary changes and have saved file."
        )
        x = input("Ready to Proceed (Y/N): ").lower()
        if x[0] == "y":
            df1 = openpyxl.load_workbook(self.file)
            ws1 = df1.active
            data1 = ws1.values
            columns1 = next(data1)[0:]
            df1 = pd.DataFrame(data1, columns=columns1)

            df2 = openpyxl.load_workbook(os.path.join(path, "old_implementations.xlsx"))
            ws2 = df2.active
            data2 = ws2.values
            columns2 = next(data2)[0:]
            df2 = pd.DataFrame(data2, columns=columns2)

            data_frame_same = df1.equals(df2)

            f = open(
                os.path.join(path, "differences.txt"),
                "w+",
            )
            f.truncate(0)

            if data_frame_same:
                self.logger.debug("No differences detected.")

            else:
                upload_data(
                    regscale_parent_id=regscale_parent_id,
                    regscale_module=regscale_module,
                )
                self.logger.debug("*** WARNING *** Differences Found.")

                # Logs changes to txt file

                diff_mask = (df1 != df2) & ~(df1.isnull() & df2.isnull())
                ne_stacked = diff_mask.stack()
                changed = ne_stacked[ne_stacked]
                changed.index.names = ["row", "col"]
                difference_locations = np.where(diff_mask)
                changed_from = df1.values[difference_locations]
                changed_to = df2.values[difference_locations]
                changes = pd.DataFrame(
                    {"from": changed_from, "to": changed_to}, index=changed.index
                )
                changes.tocsv(
                    os.path.join(path, "differences.txt"),
                    header=None,
                    index=None,
                    sep=" ",
                    mode="a",
                )
        return self.logger.info(
            "Please check differences.txt file located in artifacts folder of current working directory to see changes made."
        )

    def upload_data(self, regscale_module: str, regscale_parent_id: int) -> None:
        """
        Batch uploads updated control implementation statements to the provided RegScale parent ID
        :param str regscale_module: RegScale parent module
        :param int regscale_parent_id: RegScale parent ID
        :raises: requests.exceptions.RequestException if API call encountered an error
        :return: None
        """
        reader = pd.read_excel(self.file)
        updated_implementations = [
            {
                "id": i["Id"],
                "implementation": i["Implementation"],
                "policy": i["Policy"],
                "dateLastAssessed": None,
                "lastAssessmentResult": None,
                "controlName": i["ControlName"],
                "controlTitle": i["Description"],
                "controlId": i["ControlId"],
                "practiceLevel": None,
                "processLevel": None,
                "cyberFunction": None,
                "implementationType": None,
                "implementationMethod": None,
                "qdWellDesigned": None,
                "qdProcedures": None,
                "qdSegregation": None,
                "qdFlowdown": None,
                "qdAutomated": None,
                "qdOverall": None,
                "qiResources": None,
                "qiMaturity": None,
                "qiReporting": None,
                "qiVendorCompliance": None,
                "qiIssues": None,
                "qiOverall": None,
                "responsibility": i["Responsibility"],
                "parentID": regscale_parent_id,
                "parentModule": regscale_module,
                "inheritedControlId": None,
                "inheritedRequirementId": None,
                "inheritedSecurityPlanId": None,
                "inheritedPolicyId": None,
                "dateCreated": None,
                "lastUpdatedById": None,
                "dateLastUpdated": None,
                "weight": None,
                "isPublic": None,
                "inheritable": i["Inheritable"],
            }
            for i in reader.iterrows()
        ]
        new_implementations = json.dumps(updated_implementations)
        try:
            api.post(
                url=self.config["domain"] + "/api/controlImplementation/batchUpdate",
                json=new_implementations,
            )
            self.logger.info(
                "%s total %s for Parent ID: %s in RegScale were updated.",
                len(new_implementations),
                regscale_module,
                regscale_parent_id,
            )
            self.logger.debug("File will be deleted. Thank you.")
        except requests.exceptions.RequestException as ex:
            self.logger.error(
                "Unable to update %s for ParentId: %s in RegScale.\n%s",
                regscale_module,
                regscale_parent_id,
                ex,
            )

    def delete_file(self, path: Path):
        """
        Deletes files used during the process
        :param Path path: path to the files to delete
        :return: None
        """
        os.remove(self.file)
        os.remove(os.path.join(path, "old_implementations.xlsx"))
        os.remove(os.path.join(path, "differences.txt"))
        os.rmdir(path)
        return self.logger.debug("File has been deleted.")
