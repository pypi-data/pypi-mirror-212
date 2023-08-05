"""Flatfile-related utilities"""
import chardet
import pandas as pd
from openpyxl import Workbook

def detect_encoding(file_path):
    """Attempt to determine the encoding of a file located at the provided file path"""
    # Open the file in binary mode to prevent any decoding errors
    with open(file_path, 'rb') as f:
        # Read the first 2000 rows
        content = b''.join([f.readline() for _ in range(2000)])
        # Determine the encoding of the content
        result = chardet.detect(content)

    encoding = result['encoding']

    # If the detected encoding is ASCII, read the entire file to confirm the encoding
    if encoding.lower() == 'ascii':
        with open(file_path, 'rb') as f:
            content = f.read()
            # Determine the encoding of the entire file
            result = chardet.detect(content)
            encoding = result['encoding']

    return encoding

def analyze_dataframe(df):
    """Analyze distinct values in a dataframe"""
    analysis_dict = {}

    for col in df.columns:
        column_dict = {}

        # Add column header
        column_dict['Column'] = col
        distinct_value_count = df[col].nunique()
        distinct_value_count_string = str(format(distinct_value_count, ','))
        column_dict['Distinct Values Count'] = str(distinct_value_count)

        # Check column data type
        col_dtype = df[col].dtype
        if pd.api.types.is_string_dtype(col_dtype):
            # For string columns, calculate the maximum length
            max_size = df[col].str.len().max()
            max_size_string = str(int(max_size)) if not pd.isna(max_size) else ''
            column_dict['Max Size'] = f'Max Length: {max_size_string}'
            column_dict['Type'] = 'String'
        elif pd.api.types.is_numeric_dtype(col_dtype):
            if pd.api.types.is_integer_dtype(col_dtype):
                # For integer columns, convert the maximum value to int
                max_value = df[col].max()
                max_value_string = str(int(max_value)) if not pd.isna(max_value) else ''
                column_dict['Max Size'] = f'Max Value: {max_value_string}'
            else:
                # For other numeric columns (float), store the maximum value as is
                max_value_string = str(df[col].max())
                column_dict['Max Size'] = f'Max Value: {max_value_string}'
            column_dict['Type'] = 'Numeric'
        elif pd.api.types.is_datetime64_any_dtype(col_dtype):
            # For date or timestamp columns
            max_date = df[col].max()
            max_date_string = str(max_date) if not pd.isna(max_date) else ''
            column_dict['Max Size'] = f'Max Value: {max_date_string}'
            column_dict['Type'] = 'Date/Time'

        # Get distinct values and their counts
        col_values = df[col].copy()
        if col_values.dtype == float:
            col_values = col_values.apply(lambda x: "{:.10f}".format(x).rstrip('0').rstrip('.') if not pd.isna(x) else '<NULL>')
        col_values = col_values.fillna("<NULL>")
        value_counts = col_values.value_counts(dropna=False).replace({pd.NA: "<NULL>", pd.NaT: "<NULL>", "nan": "<NULL>"})
        if len(value_counts) > 50:
            # If there are more than 50 distinct values, store top 50 and "More than 50 distinct values"
            top_50_values = value_counts.nlargest(50).to_dict()
            # top_50_values['More than 50 distinct values'] = str(len(value_counts) - 50)
            # column_dict['Distinct Values'] = {str(k): str(format(int(v), ',')) if v != '' else str(v) for k, v in top_50_values.items()}
            column_dict['Distinct Values'] = {('{:.10f}'.format(k) if isinstance(k, float) else str(k)): str(v) if v != '' else str(v) for k, v in top_50_values.items()}
            # top_50_values['More than 50 distinct values'] = f'Distinct Values: {distinct_value_count_string}'
            column_dict['Distinct Values']['More than 50 distinct values'] = f'Distinct Values: {distinct_value_count_string}'

        else:
            # column_dict['Distinct Values'] = {str(k): str(format(int(v), ',')) if v != '' else str(v) for k, v in value_counts.items()}
            column_dict['Distinct Values'] = {('{:.10f}'.format(k) if isinstance(k, float) else str(k)): str(v) if v != '' else str(v) for k, v in value_counts.items()}

        # Add the column dictionary to the analysis dictionary
        analysis_dict[col] = column_dict

    # Create a workbook and select the active sheet
    wb = Workbook()
    ws = wb.active

    # Initialize the start column index
    start_column_index = 1

    # Iterate over each column in the analysis dictionary
    for col, col_dict in analysis_dict.items():
        # Write the "Column" and "Max Size" to the Excel file
        ws.cell(row=1, column=start_column_index, value=col_dict['Column'])
        if 'Max Size' in col_dict:
            ws.cell(row=1, column=start_column_index + 1, value=col_dict['Max Size'])

        ws.cell(row=2, column=start_column_index, value='Distinct Values')
        ws.cell(row=2, column=start_column_index + 1, value='Occurrences')

        # Write the "Distinct Values" to the Excel file
        distinct_values = col_dict['Distinct Values']
        for row_num, (key, value) in enumerate(distinct_values.items(), start=3):
            ws.cell(row=row_num, column=start_column_index, value=key)
            ws.cell(row=row_num, column=start_column_index + 1, value=value)

        # Increment the start column index for the next iteration
        start_column_index += 3

    ws.freeze_panes = 'A3'
    
    df_results = pd.DataFrame(wb.active.values)

    return df_results
