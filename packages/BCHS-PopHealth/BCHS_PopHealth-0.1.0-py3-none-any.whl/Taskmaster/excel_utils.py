from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from pandas import ExcelWriter
import win32com.client as win32
import tempfile
import os


def set_table_style(excel_writer, table_style='TableStyleMedium9'):
    worksheets = excel_writer.book.worksheets
    for i, ws in enumerate(worksheets):
        tab  = Table(displayName=f'Table{i+1}', ref=ws.dimensions)
        style = TableStyleInfo(name=table_style, showFirstColumn=False,
            showLastColumn=False, showRowStripes=True, showColumnStripes=True)

        tab.tableStyleInfo = style
        ws.add_table(tab)

def auto_fit_column_width(excel_writer):
    worksheets = excel_writer.book.worksheets
    for ws in worksheets:
        dims = {}
        for row in ws.rows:
            for cell in row:
                if cell.value:
                    dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))    
        for col, value in dims.items():
            ws.column_dimensions[col].width = value * 1.2        


def write_dataframe_to_excel(df, fpath, table_style='TableStyleMedium9', auto_fit=True):
    excel_writer = ExcelWriter(fpath, engine='openpyxl')
    df.to_excel(excel_writer, index=False)
    set_table_style(excel_writer)
    
    if auto_fit:
        auto_fit_column_width(excel_writer)
    
    excel_writer.close()

def write_dataframes_to_excel(dfs, fpath, sheet_names, table_style='TableStyleMedium9', auto_fit=True):
    excel_writer = ExcelWriter(fpath, engine='openpyxl')
    
    for i,df in enumerate(dfs):
        df.to_excel(excel_writer, sheet_name = sheet_names[i], index=False)
    set_table_style(excel_writer)

    if auto_fit:
        auto_fit_column_width(excel_writer)
    
    excel_writer.close()

def write_dataframe_to_csv(df, fpath):
    df.to_csv(fpath, index=False)

def refresh_pivot_tables(wb):
    sheet_count = wb.Sheets.Count
    
    for i in range(1, sheet_count+1):
        ws = wb.Worksheets[i]

        pivot_count = ws.PivotTables().Count
        for j in range(1, pivot_count+1):
            ws.PivotTables(j).PivotCache().Refresh()

def write_dataframes_to_excel_from_template(dfs, template_path, out_path, sheetnames=['Sheet1']):
    if type(dfs) is not list: dfs = [dfs]

    excel = win32.gencache.EnsureDispatch('Excel.Application')
    #excel.Visible = True
    excel.DisplayAlerts = False

    #Open the template file
    wb_template = excel.Workbooks.Open(template_path)

    for i, df in enumerate(dfs):
        sheetname = sheetnames[i]
        rows, cols = df.shape
        rows = rows if not df.empty else 1 #Rows can't be 0 when defining Excel Range
        tf_name = ''
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tf:
            tf_name = tf.name

        #Write dataframe contents to a temporary excel file
        df.to_excel(tf_name, index=False, header=False, sheet_name=sheetname)

        #Open the temporary data file
        wb_data = excel.Workbooks.Open(tf_name)
        
        try:
            #Copy the data from the temporary data file
            ws = wb_data.Worksheets(sheetname)
            rng_start = ws.Cells(1, 1)
            rng_end = ws.Cells(rows, cols)
            ws.Range(rng_start, rng_end).Select()
        

            #Paste the data into the template
            excel.Selection.Copy(Destination=wb_template.Worksheets(sheetname).Range('A2'))

        finally:
            wb_data.Close()
            os.remove(tf_name)

    #Refresh pivot tables and 'save as' the template
    wb_template.RefreshAll()
    wb_template.SaveAs(str(out_path))
    wb_template.Close